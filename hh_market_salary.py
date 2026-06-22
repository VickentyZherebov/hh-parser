"""
Блок 2 (ITD-1245): уровень рыночных зарплат.

ВАЖНО про метрику. Задача просит «вилку min–max», которую HH показывает в форме
создания вакансии. Без рабочего аккаунта работодателя такая вилка недоступна.
Открыто (без аккаунта) доступна только СРЕДНЯЯ зарплата со stats.hh.ru:
  - averageCompensation — средняя ПРЕДЛАГАЕМАЯ ЗП (по вакансиям, «оффер рынка»);
  - averageExpected     — средняя ОЖИДАЕМАЯ ЗП (по резюме, «запрос кандидатов»).
Это и собирает данный модуль. Расхождение «средняя vs вилка» вынесено на
согласование с постановщиком (см. план ITD-1245).

Гео задаётся списком из файла (Excel/CSV, города в первом столбце).
Результат: Excel «Город × Профобласть» — отдельный лист на предлагаемую и
на ожидаемую ЗП.

Запуск из командной строки:
    python hh_market_salary.py [путь_к_файлу_с_городами]
"""

from __future__ import annotations

import csv
import logging
import sys
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

import hh_market_stats as stats
from hh_vacancies_by_geo import resolve_city_ids

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)

# Города по умолчанию (если файл не передан) — несколько крупных для демо.
DEFAULT_CITIES = [
    "Москва", "Санкт-Петербург", "Казань", "Новосибирск", "Екатеринбург",
]

# Метрики ЗП: (ключ stats, человекочитаемое название, имя листа Excel)
SALARY_METRICS = [
    (stats.METRIC_AVG_OFFERED, "Средняя предлагаемая (вакансии)", "Предлагаемая ЗП"),
    (stats.METRIC_AVG_EXPECTED, "Средняя ожидаемая (резюме)", "Ожидаемая ЗП"),
]


def load_cities_from_file(path: str | Path) -> list[str]:
    """
    Читает список городов из файла. Поддержка .xlsx и .csv. Берётся первый
    столбец, шапка («город»/«city»/«регион») пропускается. Пустые строки
    игнорируются.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Файл со списком городов не найден: {path}")

    raw: list[str] = []
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        if not HAS_OPENPYXL:
            raise RuntimeError("Нужен openpyxl для чтения .xlsx: pip install openpyxl")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if row and row[0] is not None:
                raw.append(str(row[0]).strip())
    else:  # csv / txt
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            for cols in csv.reader(fh):
                if cols and cols[0].strip():
                    raw.append(cols[0].strip())

    # Пропускаем строку-заголовок, если она похожа на шапку
    if raw and raw[0].lower() in ("город", "city", "регион", "город/регион", "населённый пункт"):
        raw = raw[1:]

    # Дедуп с сохранением порядка
    seen, cities = set(), []
    for c in raw:
        if c and c.lower() not in seen:
            seen.add(c.lower())
            cities.append(c)
    return cities


def build_salary_table(
    cities: Optional[list[str]] = None,
    prof_area_ids: Optional[list] = None,
    *,
    cache_path: Optional[str] = None,
    progress_cb: Optional[Callable[[dict], None]] = None,
) -> dict:
    """
    Собирает таблицу средних ЗП «город × профобласть» (последнее значение) для
    обеих метрик (предлагаемая и ожидаемая).

    Возвращает dict, аналогичный build_index_table, но в каждой строке есть
    поля по обеим метрикам: {"averageCompensation": value|None,
    "averageExpected": value|None} + период.
    """
    cities = cities or DEFAULT_CITIES

    data = stats.fetch_stats_data(cache_path=cache_path)
    region_ids = stats.available_region_ids(data)
    name_map = stats.prof_area_name_map(include_all=True)

    if prof_area_ids is None:
        prof_cols = stats.load_prof_areas(include_all=True)
    else:
        prof_cols = [{"id": pid, "name": name_map.get(str(pid), f"Профобласть {pid}")}
                     for pid in prof_area_ids]

    resolved = resolve_city_ids(cities)
    missing_cities = [c for c in cities if c not in resolved]
    areas_full = stats.load_areas_full()

    present: dict[str, dict] = {}
    region_fallback: list[str] = []
    missing_in_stats: list[str] = []
    for city, area_id in resolved.items():
        res = stats.resolve_to_stats_region(area_id, region_ids, areas_full=areas_full)
        if res is None:
            missing_in_stats.append(city)
            continue
        present[city] = {"area_id": res["id"], "level": res["level"], "stats_name": res["name"]}
        if res["level"] == "region":
            region_fallback.append(f'{city} → {res["name"]}')

    rows: list[dict] = []
    period_counter: dict[str, int] = {}
    total = len(present) * len(prof_cols)
    done = 0

    for city, meta in present.items():
        area_id = meta["area_id"]
        for pa in prof_cols:
            row = {
                "city": city,
                "area_id": area_id,
                "level": meta["level"],
                "stats_name": meta["stats_name"],
                "prof_area_id": pa["id"],
                "prof_area_name": pa["name"],
            }
            for metric_key, _, _ in SALARY_METRICS:
                lv = stats.get_last(data, area_id, metric_key, pa["id"])
                row[metric_key] = round(lv["value"]) if lv else None
                if lv:
                    key = f'{lv["month_name"]} {lv["year"]}'
                    period_counter[key] = period_counter.get(key, 0) + 1
            rows.append(row)
            done += 1
            if progress_cb:
                progress_cb({"done": done, "total": total,
                             "status_text": f'{city} — {pa["name"]}'})

    as_of = max(period_counter, key=period_counter.get) if period_counter else "—"

    return {
        "scraped_at": datetime.now().isoformat(timespec="seconds"),
        "as_of": as_of,
        "cities": list(present.keys()),
        "prof_areas": prof_cols,
        "rows": rows,
        "region_fallback": region_fallback,
        "missing_cities": missing_cities,
        "missing_in_stats": missing_in_stats,
    }


def export_salary_to_xlsx(result: dict, output_path: str) -> None:
    """
    Экспорт build_salary_table() в Excel. По листу на каждую метрику ЗП
    (предлагаемая / ожидаемая): сводная «Город × Профобласть», значения в рублях.
    Над таблицей — пометка про период, источник и «это средняя, не вилка».
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("Нужен openpyxl: pip install openpyxl")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="548235")
    note_font = Font(italic=True, color="808080")
    warn_font = Font(italic=True, color="C00000")
    city_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    prof_areas = result["prof_areas"]
    cities = result["cities"]

    wb = openpyxl.Workbook()

    for sheet_i, (metric_key, metric_title, sheet_name) in enumerate(SALARY_METRICS):
        ws = wb.active if sheet_i == 0 else wb.create_sheet()
        ws.title = sheet_name

        ws.append([f"Рыночные зарплаты: {metric_title}"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=13)
        ws.append([f"Источник: stats.hh.ru · данные на: {result.get('as_of', '—')} · "
                   f"выгружено: {result.get('scraped_at', '')}"])
        ws.cell(ws.max_row, 1).font = note_font
        ws.append(["⚠ Это СРЕДНЕЕ значение, а не вилка min–max. Вилку HH показывает "
                   "только в форме создания вакансии (нужен аккаунт работодателя)."])
        ws.cell(ws.max_row, 1).font = warn_font
        ws.append([])

        header = ["Город / Регион"] + [pa["name"] for pa in prof_areas]
        ws.append(header)
        hr = ws.max_row
        for cell in ws[hr]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        idx = {(r["city"], str(r["prof_area_id"])): r.get(metric_key) for r in result["rows"]}
        level_by_city = {r["city"]: r.get("level", "city") for r in result["rows"]}

        for city in cities:
            label = f"{city} *" if level_by_city.get(city) == "region" else city
            line = [label]
            for pa in prof_areas:
                v = idx.get((city, str(pa["id"])))
                line.append(v if v is not None else "—")
            ws.append(line)
            ws.cell(ws.max_row, 1).font = city_font
            for c in range(2, len(header) + 1):
                cell = ws.cell(ws.max_row, c)
                cell.alignment = center
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "# ##0"

        if result.get("region_fallback") or result.get("missing_cities") or result.get("missing_in_stats"):
            ws.append([])
            if result.get("region_fallback"):
                ws.append(["* данные уровня региона (города нет в статистике отдельно): "
                           + "; ".join(result["region_fallback"])])
                ws.cell(ws.max_row, 1).font = note_font
            if result.get("missing_cities"):
                ws.append([f"Не найдены в справочнике HH: {', '.join(result['missing_cities'])}"])
                ws.cell(ws.max_row, 1).font = note_font
            if result.get("missing_in_stats"):
                ws.append([f"Нет в статистике stats.hh.ru: {', '.join(result['missing_in_stats'])}"])
                ws.cell(ws.max_row, 1).font = note_font

        ws.column_dimensions["A"].width = 22
        for i in range(2, len(header) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 16
        ws.freeze_panes = ws.cell(row=hr + 1, column=2)

    wb.save(output_path)
    logger.info("Сохранено: %s", output_path)


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s",
                        datefmt="%H:%M:%S")
    cache = str(Path(__file__).parent / "data" / "stats_hh_cache.json")

    cities = None
    if len(sys.argv) > 1:
        cities = load_cities_from_file(sys.argv[1])
        print(f"Городов из файла {sys.argv[1]}: {len(cities)}")

    print("=" * 60)
    print("Блок 2: рыночные зарплаты (средняя предлагаемая + ожидаемая)")
    print("=" * 60)

    result = build_salary_table(
        cities=cities,
        cache_path=cache,
        progress_cb=lambda p: print(f"  [{p['done']}/{p['total']}] {p['status_text']}", flush=True),
    )

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out = str(Path(__file__).parent / f"hh_salary_{ts}.xlsx")
    export_salary_to_xlsx(result, out)
    print(f"\nГотово: {out}")
    print(f"Города: {len(result['cities'])}, профобласти: {len(result['prof_areas'])}, "
          f"данные на: {result['as_of']}")
    if result["missing_in_stats"]:
        print(f"Нет в статистике: {', '.join(result['missing_in_stats'])}")
    if result["missing_cities"]:
        print(f"Не найдены в HH: {', '.join(result['missing_cities'])}")


if __name__ == "__main__":
    main()
