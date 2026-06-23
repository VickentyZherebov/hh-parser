"""
Додо гео-парсер для веб-морды (вкладка «Додо»): сбор из открытого API воронки
HH (hh.htdev.ru/dodo_api.php), запись в прод-БД авилидс (через сайдкар-туннель
avileads-db-tunnel) и сборка гео-Excel «куда постить».

Логика «открыто/закрыто» = точная копия фильтра HH-страницы (funnel_open).
Каноничная CLI-версия — work/avileads/apps/dodo-geo-parser/ (для крона);
здесь self-contained копия для on-demand кнопки Маши.
"""
from __future__ import annotations

import io
import os
import time
from concurrent.futures import ThreadPoolExecutor

import requests

API = "https://hh.htdev.ru/dodo_api.php"
KEY = os.environ.get("DODO_API_KEY", "jgn3w9ufhw3fwq3f3wbffiowsegsd32q4r5q2")
DSN = os.environ.get("DODO_PG_DSN", "")  # postgresql://claude_ai:...@avileads-db-tunnel:5432/avileads

BASE_ROLES = ["Кассир", "Пиццамейкер", "Пеший курьер", "Курьер на личном автомобиле"]
VAC_FIELDS = {
    "salary": "salary", "salaryAfterTax": "salary_after_tax", "hourlyRate": "hourly_rate",
    "monthlyWorkingHours": "monthly_working_hours", "hasBonus": "has_bonus",
    "hourlyBonus": "hourly_bonus", "otherBonus": "other_bonus", "ratePerOrder": "rate_per_order",
    "ratePerTrip": "rate_per_trip", "amortizationPayment": "amortization_payment",
    "hasFuelReimbursement": "fuel_reimbursement",
}
VAC_COLS = list(VAC_FIELDS.values())


def funnel_open(name: str) -> bool:
    """Точная копия фильтра JS HH-страницы (checkPermittedName)."""
    n = (name or "").lower()
    if any(x in n for x in ("велосипед", "скутер", "мопед")):
        return False
    return any(x in n for x in ("курьер", "кассир", "пиццамейкер"))


def _call(task: str, *, retries: int = 6, **extra) -> list:
    files = {k: (None, v) for k, v in {"special": KEY, "task": task,
             **{k: str(v) for k, v in extra.items()}}.items()}
    last = None
    for attempt in range(retries):
        try:
            r = requests.post(API, files=files, timeout=30)
            if r.status_code in (502, 503, 504, 429):
                raise RuntimeError(f"HTTP {r.status_code}")
            r.raise_for_status()
            j = r.json()
            if j.get("status") != 1:
                raise RuntimeError(j.get("message"))
            return j.get("data", []) or []
        except Exception as e:  # noqa: BLE001
            last = e
            time.sleep(min(2.0 * (attempt + 1), 12))
    raise RuntimeError(f"{task} {extra}: {last}")


def run(progress_cb=None, *, workers: int = 5) -> dict:
    """Сбор + запись в прод-БД (fresh, атомарно). Возвращает статистику."""
    import psycopg

    def log(msg):
        if progress_cb:
            progress_cb(msg)

    # ── сбор ──
    cities = []
    for _ in range(4):
        cities = _call("getcities")
        if cities:
            break
        time.sleep(2)
    if not cities:
        raise RuntimeError("getcities пуст — источник недоступен")
    log(f"Городов: {len(cities)}")

    units = []
    def _units(city):
        try:
            rs = _call("getrestaurants", localityId=city["id"])
        except Exception:
            return []
        return [{"city_uuid": city["id"], "uuid": u["id"], "seq": i,
                 "address": (u.get("address") or "").strip(),
                 "name": (u.get("name") or "").strip()}
                for i, u in enumerate(rs, 1)]
    with ThreadPoolExecutor(max_workers=workers) as ex:
        for ch in ex.map(_units, cities):
            units.extend(ch)
    log(f"Адресов: {len(units)}")

    vac_map = {}
    def _vac(u):
        try:
            return u["uuid"], _call("getvacancies", unitUuid=u["uuid"])
        except Exception:
            return u["uuid"], None
    with ThreadPoolExecutor(max_workers=workers) as ex:
        for uuid, v in ex.map(_vac, units):
            if v is not None:
                vac_map[uuid] = v
    log(f"Адресов с ответом по вакансиям: {len(vac_map)}")

    # ── запись (fresh, в одной транзакции) ──
    log("Запись в БД…")
    with psycopg.connect(DSN, autocommit=False) as conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM parser_geo_dodo_vacancy")
        cur.execute("DELETE FROM parser_geo_dodo_address")
        cur.execute("DELETE FROM parser_geo_dodo_city")
        cur.executemany(
            "INSERT INTO parser_geo_dodo_city (name,dodo_city_uuid,status,updated_at) "
            "VALUES (%s,%s,'processed',now()) ON CONFLICT (name) DO UPDATE SET "
            "dodo_city_uuid=EXCLUDED.dodo_city_uuid,status='processed',updated_at=now()",
            [(c["name"], c["id"]) for c in cities])
        cur.execute("SELECT dodo_city_uuid,id FROM parser_geo_dodo_city WHERE dodo_city_uuid IS NOT NULL")
        city_id = dict(cur.fetchall())
        cur.executemany(
            "INSERT INTO parser_geo_dodo_address "
            "(city_id,address_name,address_seq_number,dodo_unit_uuid,address_full,status,updated_at) "
            "VALUES (%s,%s,%s,%s,%s,'processed',now()) ON CONFLICT (dodo_unit_uuid) DO UPDATE SET "
            "city_id=EXCLUDED.city_id,address_name=EXCLUDED.address_name,"
            "address_seq_number=EXCLUDED.address_seq_number,address_full=EXCLUDED.address_full,"
            "status='processed',updated_at=now()",
            [(city_id[u["city_uuid"]], u["address"], u["seq"], u["uuid"],
              f'{u["name"]} — {u["address"]}'.strip(" —"))
             for u in units if u["city_uuid"] in city_id])
        cur.execute("SELECT dodo_unit_uuid,id FROM parser_geo_dodo_address WHERE dodo_unit_uuid IS NOT NULL")
        addr_id = dict(cur.fetchall())

        set_clause = ", ".join(f"{c}=EXCLUDED.{c}" for c in (VAC_COLS + ["status", "dodo_vacancy_uuid"]))
        sql = ("INSERT INTO parser_geo_dodo_vacancy "
               "(address_id,vacancy_type,status,dodo_vacancy_uuid," + ",".join(VAC_COLS) + ",updated_at) "
               "VALUES (" + ",".join(["%s"] * (4 + len(VAC_COLS))) + ",now()) "
               "ON CONFLICT (address_id,vacancy_type) DO UPDATE SET " + set_clause + ",updated_at=now()")
        rows = []
        def vals(aid, vt, st, r):
            return tuple([aid, vt, st, (r or {}).get("id")] + [(r or {}).get(k) for k in VAC_FIELDS])
        for u in units:
            aid = addr_id.get(u["uuid"])
            if aid is None:
                continue
            by = {r["name"]: r for r in vac_map.get(u["uuid"], [])}
            for name, r in by.items():
                rows.append(vals(aid, name, "enabled" if funnel_open(name) else "ratecard", r))
            for vt in BASE_ROLES:
                if vt not in by:
                    rows.append(vals(aid, vt, "disabled", None))
        cur.executemany(sql, rows)
        conn.commit()

    stats = {"cities": len(cities), "addresses": len(units), "vacancies": len(rows)}
    log(f"Готово: {stats['cities']} городов, {len(addr_id)} пиццерий, {stats['vacancies']} строк")
    return stats


def build_geo_xlsx() -> bytes:
    """Гео-Excel «куда постить» из текущей БД. Возвращает байты файла."""
    import psycopg
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from datetime import datetime

    with psycopg.connect(DSN) as conn:
        cur = conn.cursor()
        pivot = cur.execute(
            "SELECT c.name,v.vacancy_type,count(*) FROM parser_geo_dodo_vacancy v "
            "JOIN parser_geo_dodo_address a ON a.id=v.address_id "
            "JOIN parser_geo_dodo_city c ON c.id=a.city_id "
            "WHERE v.status='enabled' GROUP BY c.name,v.vacancy_type").fetchall()
        units = dict(cur.execute(
            "SELECT c.name,count(*) FROM parser_geo_dodo_address a "
            "JOIN parser_geo_dodo_city c ON c.id=a.city_id GROUP BY c.name").fetchall())
        detail = cur.execute(
            "SELECT c.name,a.address_name,v.vacancy_type,v.salary,v.hourly_rate "
            "FROM parser_geo_dodo_vacancy v JOIN parser_geo_dodo_address a ON a.id=v.address_id "
            "JOIN parser_geo_dodo_city c ON c.id=a.city_id WHERE v.status='enabled' "
            "ORDER BY c.name,a.address_name,v.vacancy_type").fetchall()

    cell, role_tot, cities = {}, {}, set()
    for city, role, n in pivot:
        cell[(city, role)] = n
        role_tot[role] = role_tot.get(role, 0) + n
        cities.add(city)
    roles = sorted(role_tot, key=role_tot.get, reverse=True)
    city_tot = {c: sum(cell.get((c, r), 0) for r in roles) for c in cities}
    cities_sorted = sorted(cities, key=lambda c: city_tot[c], reverse=True)

    HF, HFILL = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="C00000")
    CEN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Гео × роли"
    ws.append([f"Додо — открытые вакансии по гео (куда постить) · {datetime.now():%d.%m.%Y %H:%M}"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=13)
    ws.append([])
    header = ["Город", "Точек в городе", "Открытых всего"] + roles
    ws.append(header)
    hr = ws.max_row
    for c in ws[hr]:
        c.font, c.fill, c.alignment = HF, HFILL, CEN
    for city in cities_sorted:
        ws.append([city, units.get(city, 0), city_tot[city]] + [cell.get((city, r), "") for r in roles])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        for i in range(2, len(header) + 1):
            ws.cell(ws.max_row, i).alignment = CEN
    ws.freeze_panes = ws.cell(row=hr + 1, column=2)
    ws.column_dimensions["A"].width = 22
    for i in range(2, len(header) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14

    ws2 = wb.create_sheet("Детально")
    ws2.append(["Город", "Адрес", "Роль", "ЗП, ₽", "₽/час"])
    for c in ws2[1]:
        c.font, c.fill, c.alignment = HF, HFILL, CEN
    for r in detail:
        ws2.append(list(r))
    ws2.freeze_panes = "A2"
    for col, w in zip("ABCDE", (20, 34, 30, 12, 10)):
        ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
