"""
Парсинг вакансий HH.ru по гео и профессиональным ролям.

Два режима поиска:
1. По названиям вакансий (текстовый поиск по полю name)
2. По профессиональным ролям (professional_role ID из справочника HH.ru)

Для каждой комбинации город × запрос считаем:
- Общее кол-во вакансий (found из API)
- Уникальных работодателей (distinct employer_id по пагинации, макс 2000)
- Ссылка на поиск HH.ru для ручной проверки

Результат: структурированный dict + экспорт в Excel (два листа + сводные таблицы).

Использование как модуль:
    from hh_vacancies_by_geo import (
        load_area_tree, load_professional_roles,
        resolve_city_ids, scrape_vacancies_geo, export_to_xlsx,
    )

Использование из командной строки:
    python hh_vacancies_by_geo.py
"""

import logging
import time
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional
from urllib.parse import urlencode

import requests

from hh_browser_fetcher import CellResult, CellTask, scrape_via_browser

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

__version__ = "1.0.0"

logger = logging.getLogger(__name__)

BASE_API_URL = "https://api.hh.ru"
HH_SEARCH_URL = "https://hh.ru/search/vacancy"
HEADERS = {"User-Agent": "HH-Parser-Web/1.0"}

# Задержка между запросами (rate limiting)
REQUEST_DELAY = 0.25

# HH.ru отдаёт максимум 2000 результатов (20 страниц по 100)
MAX_PAGES = 20
PAGE_SIZE = 100

# ── Кэши (загружаются один раз за процесс) ────────────────────────────────────
_area_tree_cache: Optional[dict] = None       # city_name.lower() -> (area_id, canonical_name)
_professional_roles_cache: Optional[list] = None  # list of {id, name}


# ── Загрузка справочников ─────────────────────────────────────────────────────

def load_area_tree() -> dict:
    """
    Загружает полное дерево регионов HH.ru (/areas) и строит плоский индекс:
        city_name.lower() -> (area_id: int, canonical_name: str)

    Результат кэшируется в модульной переменной — повторный вызов возвращает кэш.
    """
    global _area_tree_cache
    if _area_tree_cache is not None:
        return _area_tree_cache

    logger.info("Загружаю дерево регионов HH.ru...")
    resp = requests.get(f"{BASE_API_URL}/areas", headers=HEADERS, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    index: dict[str, tuple[int, str]] = {}

    def _walk(node: dict) -> None:
        name = node.get("name", "").strip()
        area_id = int(node["id"])
        if name:
            index[name.lower()] = (area_id, name)
        for child in node.get("areas", []):
            _walk(child)

    for country in data:
        _walk(country)

    logger.info("Загружено %d регионов/городов", len(index))
    _area_tree_cache = index
    return index


def load_professional_roles() -> list[dict]:
    """
    Загружает справочник профессиональных ролей HH.ru (/professional_roles).
    Возвращает плоский список: [{"id": int, "name": str}, ...]

    Результат кэшируется в модульной переменной.
    """
    global _professional_roles_cache
    if _professional_roles_cache is not None:
        return _professional_roles_cache

    logger.info("Загружаю справочник профессиональных ролей HH.ru...")
    resp = requests.get(f"{BASE_API_URL}/professional_roles", headers=HEADERS, timeout=30)
    resp.raise_for_status()
    data = resp.json()

    roles: list[dict] = []
    seen_ids: set[int] = set()

    # Структура: {"categories": [{"id": ..., "name": ..., "roles": [{...}]}]}
    # Роль может встречаться в нескольких категориях — дедуплицируем по ID
    for category in data.get("categories", []):
        for role in category.get("roles", []):
            role_id = role.get("id")
            role_name = role.get("name", "").strip()
            if role_id and role_name and role_id not in seen_ids:
                seen_ids.add(role_id)
                roles.append({"id": int(role_id), "name": role_name})

    roles.sort(key=lambda r: r["name"].lower())
    logger.info("Загружено %d профессиональных ролей", len(roles))
    _professional_roles_cache = roles
    return roles


# ── Резолв городов ────────────────────────────────────────────────────────────

def resolve_city_ids(city_names: list[str]) -> dict[str, int]:
    """
    Определяет area_id HH.ru для каждого города по названию.

    Загружает дерево регионов (кэшируется) и ищет по точному совпадению
    нижнего регистра. Города, которые не удалось найти, пропускаются
    (в лог пишется предупреждение).

    Возвращает dict: city_name -> area_id (только найденные города).
    """
    area_index = load_area_tree()
    result: dict[str, int] = {}
    not_found: list[str] = []

    for city in city_names:
        key = city.lower().strip()
        if key in area_index:
            area_id, canonical = area_index[key]
            result[city] = area_id
            logger.debug("Город найден: %s -> area_id=%d (%s)", city, area_id, canonical)
        else:
            not_found.append(city)
            logger.warning("Город не найден в дереве HH.ru: %s", city)

    if not_found:
        logger.warning("Не найдены города: %s", ", ".join(not_found))

    return result


# ── Формирование URL для проверки ─────────────────────────────────────────────

def make_search_url(
    area_id: int,
    text: Optional[str] = None,
    professional_role: Optional[int] = None,
) -> str:
    """
    Формирует ссылку на поиск вакансий на HH.ru для ручной проверки в браузере.

    Параметры:
        area_id          — ID региона/города HH.ru
        text             — текстовый запрос (поиск по названию вакансии)
        professional_role — ID профессиональной роли из справочника HH.ru
    """
    params: dict = {"area": area_id}
    if text:
        params["text"] = text
        params["search_field"] = "name"
    if professional_role is not None:
        params["professional_role"] = professional_role
    return f"{HH_SEARCH_URL}?{urlencode(params)}"


# ── Запрос к API ──────────────────────────────────────────────────────────────

def fetch_vacancies(
    area_id: int,
    text: Optional[str] = None,
    professional_role: Optional[int] = None,
    filter_text: Optional[str] = None,
    filter_exact: bool = False,
) -> tuple[int, int]:
    """
    Запрашивает вакансии через HH.ru API и возвращает (total_count, unique_employer_count).

    total_count          — общее кол-во вакансий (поле "found" из первого ответа API)
    unique_employer_count — кол-во уникальных работодателей (distinct employer_id
                           по всем страницам пагинации, максимум 2000 результатов)

    Если filter_text задан — вакансии фильтруются на нашей стороне,
    потому что HH API возвращает нерелевантные результаты даже с кавычками.
    - filter_exact=False (частичное): название должно СОДЕРЖАТЬ фразу
    - filter_exact=True  (точное):   название должно ТОЧНО СОВПАДАТЬ с фразой

    Параметры:
        area_id           — ID региона/города HH.ru
        text              — текстовый поиск по названию вакансии
        professional_role — фильтр по профессиональной роли (ID из справочника)
        filter_text       — пост-фильтр по названию вакансии
        filter_exact      — True = точное совпадение, False = вхождение фразы
    """
    params: dict = {
        "area": area_id,
        "per_page": PAGE_SIZE,
        "page": 0,
    }
    if text:
        params["text"] = text
        params["search_field"] = "name"
    if professional_role is not None:
        params["professional_role"] = professional_role

    employer_ids: set[str] = set()
    total = 0
    filtered_total = 0
    need_filter = filter_text is not None
    filter_lower = filter_text.lower() if need_filter else ""

    for page in range(MAX_PAGES):
        params["page"] = page

        resp = requests.get(
            f"{BASE_API_URL}/vacancies",
            params=params,
            headers=HEADERS,
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()

        if page == 0:
            total = data.get("found", 0)

        items = data.get("items", [])
        if not items:
            break

        for item in items:
            if need_filter:
                name = (item.get("name") or "").lower()
                if filter_exact:
                    if name != filter_lower:
                        continue
                else:
                    if filter_lower not in name:
                        continue
                filtered_total += 1
            employer = item.get("employer") or {}
            emp_id = employer.get("id")
            if emp_id:
                employer_ids.add(str(emp_id))

        pages_available = data.get("pages", 0)
        if page + 1 >= pages_available:
            break

        time.sleep(REQUEST_DELAY)

    final_total = filtered_total if need_filter else total
    return final_total, len(employer_ids)


# ── Основная функция парсинга ─────────────────────────────────────────────────

def scrape_vacancies_geo(
    city_ids: dict[str, int],
    role_ids: list[int],
    search_texts: list,
    progress_cb: Optional[Callable[[dict], None]] = None,
    pool_size: int = 10,
) -> dict:
    """
    Главная функция: обходит все комбинации город × запрос через пул
    Playwright-контекстов с ротацией прокси (каждый воркер — sticky прокси,
    ротация контекста каждые 30 задач или 5 минут).

    Сигнатура и формат возврата сохранены — UI/Excel работают как раньше.
    Изменился только бэкенд: вместо api.hh.ru/vacancies (закрыт HH) теперь
    парсим публичный hh.ru/search/vacancy через настоящий браузер.

    Аргументы:
        city_ids     — dict: city_name -> area_id (результат resolve_city_ids)
        role_ids     — список ID профессиональных ролей HH.ru
        search_texts — список текстовых запросов: строки или dict
                       {"text": str, "exact": bool}
        progress_cb  — опциональный коллбэк прогресса, вызывается с dict:
                       {done, total, city, query, status_text}
        pool_size    — сколько параллельных контекстов держим (по умолчанию 10)
    """
    # ── Имена ролей для отображения ───────────────────────────────────────────
    role_names_map: dict[int, str] = {}
    if role_ids:
        try:
            all_roles = load_professional_roles()
            role_names_map = {r["id"]: r["name"] for r in all_roles}
        except Exception as exc:
            logger.warning("Не удалось загрузить справочник ролей: %s", exc)

    # ── Нормализация search_texts ─────────────────────────────────────────────
    def _normalize_search(item) -> tuple[str, bool]:
        if isinstance(item, str):
            return item, False
        return item.get("text", ""), item.get("exact", False)

    norm_texts: list[tuple[str, bool]] = [
        ns for ns in (_normalize_search(it) for it in search_texts) if ns[0]
    ]

    # ── Сборка списка задач ───────────────────────────────────────────────────
    # Порядок: сначала роли, потом тексты. Внутри — по порядку городов.
    # Сохраняем индексы, чтобы потом разложить результаты обратно.
    tasks: list[CellTask] = []
    role_task_idx: dict[tuple[str, int], int] = {}
    text_task_idx: dict[tuple[str, str, bool], int] = {}

    for city, area_id in city_ids.items():
        for role_id in role_ids:
            tasks.append(CellTask(
                city=city,
                area_id=area_id,
                professional_role=role_id,
                role_name=role_names_map.get(role_id, f"Роль #{role_id}"),
            ))
            role_task_idx[(city, role_id)] = len(tasks) - 1

    for city, area_id in city_ids.items():
        for search_text, exact in norm_texts:
            api_text = f'"{search_text}"' if exact else search_text
            tasks.append(CellTask(
                city=city,
                area_id=area_id,
                text=api_text,
                filter_text=search_text,
                filter_exact=exact,
            ))
            text_task_idx[(city, search_text, exact)] = len(tasks) - 1

    # ── Адаптер progress_cb из формата BrowserPool в формат UI ────────────────
    def _adapter(event: dict) -> None:
        if progress_cb is None:
            return
        et = event.get("type")
        if et == "cell_done":
            status = (
                f"[{event['done']}/{event['total']}] "
                f"{event['city']} — {event['query']}: "
                f"всего={event['total_vacancies']}, "
                f"уник={event['unique_employers']}"
            )
        elif et == "cell_error":
            status = (
                f"[{event['done']}/{event['total']}] "
                f"{event['city']} — {event['query']}: "
                f"ошибка ({event.get('error', '?')})"
            )
        elif et == "context_open":
            # Не шлём в UI — это шум
            return
        else:
            return
        try:
            progress_cb({
                "done": event["done"],
                "total": event["total"],
                "city": event["city"],
                "query": event["query"],
                "status_text": status,
            })
        except Exception:
            pass

    # ── Запуск пула ───────────────────────────────────────────────────────────
    results = scrape_via_browser(tasks, pool_size=pool_size, progress_cb=_adapter)

    # Привязываем результат к индексу задачи через identity (CellResult.task is task)
    result_by_idx: dict[int, CellResult] = {}
    for r in results:
        for idx, t in enumerate(tasks):
            if r.task is t:
                result_by_idx[idx] = r
                break

    # ── Раскладываем результаты обратно в roles[] и texts[] ───────────────────
    roles_results: list[dict] = []
    texts_results: list[dict] = []

    for city, area_id in city_ids.items():
        for role_id in role_ids:
            idx = role_task_idx[(city, role_id)]
            r = result_by_idx.get(idx)
            role_name = role_names_map.get(role_id, f"Роль #{role_id}")
            url = make_search_url(area_id, professional_role=role_id)
            if r and not r.error:
                total = r.total_vacancies
                unique = r.unique_employers
            else:
                total, unique = -1, -1
            roles_results.append({
                "city": city,
                "role_id": role_id,
                "role_name": role_name,
                "total": total,
                "unique": unique,
                "url": url,
            })

    for city, area_id in city_ids.items():
        for search_text, exact in norm_texts:
            idx = text_task_idx[(city, search_text, exact)]
            r = result_by_idx.get(idx)
            api_text = f'"{search_text}"' if exact else search_text
            url = make_search_url(area_id, text=api_text)
            if r and not r.error:
                # Для текстовых запросов используем matched_vacancies
                # (после пост-фильтра по filter_text/filter_exact)
                total = r.matched_vacancies
                unique = r.unique_employers
            else:
                total, unique = -1, -1
            texts_results.append({
                "city": city,
                "text": search_text,
                "exact": exact,
                "total": total,
                "unique": unique,
                "url": url,
            })

    return {
        "scraped_at": datetime.now().isoformat(timespec="seconds"),
        "cities": list(city_ids.keys()),
        "roles": roles_results,
        "texts": texts_results,
    }


# ── Экспорт в Excel ───────────────────────────────────────────────────────────

def export_to_xlsx(result: dict, output_path: str) -> None:
    """
    Экспортирует результат scrape_vacancies_geo() в Excel-файл.

    Структура файла:
    - Лист "По ролям":    сырые данные + сводная таблица (город × роль, уникальные)
    - Лист "По названиям": сырые данные + сводная таблица (город × запрос, уникальные)

    В сырых данных столбец "Ссылка" содержит кликабельную гиперссылку "проверить".
    Заголовки выделены синим фоном и жирным шрифтом.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError(
            "Библиотека openpyxl не установлена. "
            "Выполните: pip install openpyxl"
        )

    # ── Стили ─────────────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    pivot_header_font = Font(bold=True)
    pivot_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=12)
    link_font = Font(color="0563C1", underline="single")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _apply_header_row(ws, row_num: int, fill=None, font=None) -> None:
        f = fill or header_fill
        fo = font or header_font
        for cell in ws[row_num]:
            cell.font = fo
            cell.fill = f
            cell.alignment = center_align

    def _autowidth(ws, max_width: int = 50) -> None:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, max_width)

    def _write_link_cell(ws, row: int, col: int, url: str) -> None:
        cell = ws.cell(row=row, column=col)
        if url:
            cell.value = "проверить"
            cell.hyperlink = url
            cell.font = link_font
        else:
            cell.value = ""

    wb = openpyxl.Workbook()

    cities_order: list[str] = result.get("cities", [])

    # ══════════════════════════════════════════════════════════════════════════
    # Лист 1: По ролям
    # ══════════════════════════════════════════════════════════════════════════
    ws_roles = wb.active
    ws_roles.title = "По ролям"

    roles_data: list[dict] = result.get("roles", [])

    # ── Сырые данные ──────────────────────────────────────────────────────────
    raw_headers_roles = ["Город", "ID роли", "Роль", "Всего вакансий", "Уникальных работодателей", "Ссылка"]
    ws_roles.append(raw_headers_roles)
    _apply_header_row(ws_roles, ws_roles.max_row)

    for row in roles_data:
        r_num = ws_roles.max_row + 1
        ws_roles.append([
            row["city"],
            row["role_id"],
            row["role_name"],
            row["total"] if row["total"] >= 0 else "ошибка",
            row["unique"] if row["unique"] >= 0 else "ошибка",
            "",  # ссылка — отдельно
        ])
        _write_link_cell(ws_roles, r_num, 6, row.get("url", ""))

    # ── Разделитель ───────────────────────────────────────────────────────────
    ws_roles.append([])
    ws_roles.append(["СВОДНАЯ: уникальных работодателей (город × роль)"])
    ws_roles.cell(ws_roles.max_row, 1).font = section_font

    # ── Сводная таблица (город × роль, уникальные) ────────────────────────────
    # Собираем упорядоченный список ролей (сохраняем порядок первого появления)
    seen_role_ids: list[tuple[int, str]] = []
    seen_role_set: set[int] = set()
    for row in roles_data:
        rid = row["role_id"]
        if rid not in seen_role_set:
            seen_role_ids.append((rid, row["role_name"]))
            seen_role_set.add(rid)

    pivot_header_roles = ["Город"] + [name for _, name in seen_role_ids]
    ws_roles.append(pivot_header_roles)
    _apply_header_row(ws_roles, ws_roles.max_row, fill=pivot_header_fill, font=pivot_header_font)

    # Индекс для быстрого поиска: (city, role_id) -> unique
    roles_index: dict[tuple[str, int], int] = {
        (r["city"], r["role_id"]): r["unique"]
        for r in roles_data
        if r["unique"] >= 0
    }

    for city in cities_order:
        pivot_row = [city] + [
            roles_index.get((city, rid), 0)
            for rid, _ in seen_role_ids
        ]
        ws_roles.append(pivot_row)

    _autowidth(ws_roles)

    # ══════════════════════════════════════════════════════════════════════════
    # Лист 2: По названиям
    # ══════════════════════════════════════════════════════════════════════════
    ws_texts = wb.create_sheet("По названиям")

    texts_data: list[dict] = result.get("texts", [])

    # ── Сырые данные ──────────────────────────────────────────────────────────
    raw_headers_texts = ["Город", "Поисковый запрос", "Всего вакансий", "Уникальных работодателей", "Ссылка"]
    ws_texts.append(raw_headers_texts)
    _apply_header_row(ws_texts, ws_texts.max_row)

    for row in texts_data:
        r_num = ws_texts.max_row + 1
        ws_texts.append([
            row["city"],
            row["text"],
            row["total"] if row["total"] >= 0 else "ошибка",
            row["unique"] if row["unique"] >= 0 else "ошибка",
            "",  # ссылка — отдельно
        ])
        _write_link_cell(ws_texts, r_num, 5, row.get("url", ""))

    # ── Разделитель ───────────────────────────────────────────────────────────
    ws_texts.append([])
    ws_texts.append(["СВОДНАЯ: уникальных работодателей (город × запрос)"])
    ws_texts.cell(ws_texts.max_row, 1).font = section_font

    # ── Сводная таблица (город × текст, уникальные) ───────────────────────────
    seen_texts: list[str] = []
    seen_texts_set: set[str] = set()
    for row in texts_data:
        t = row["text"]
        if t not in seen_texts_set:
            seen_texts.append(t)
            seen_texts_set.add(t)

    pivot_header_texts = ["Город"] + seen_texts
    ws_texts.append(pivot_header_texts)
    _apply_header_row(ws_texts, ws_texts.max_row, fill=pivot_header_fill, font=pivot_header_font)

    texts_index: dict[tuple[str, str], int] = {
        (r["city"], r["text"]): r["unique"]
        for r in texts_data
        if r["unique"] >= 0
    }

    for city in cities_order:
        pivot_row = [city] + [
            texts_index.get((city, t), 0)
            for t in seen_texts
        ]
        ws_texts.append(pivot_row)

    _autowidth(ws_texts)

    wb.save(output_path)
    logger.info("Результаты сохранены: %s", output_path)


# ── Standalone-запуск ─────────────────────────────────────────────────────────

# Тестовые данные для запуска из командной строки
_DEFAULT_CITIES = [
    "Архангельск", "Белгород", "Брянск", "Владивосток", "Воронеж",
    "Елец", "Кемерово", "Комсомольск-на-Амуре", "Курск", "Липецк",
    "Омск", "Оренбург", "Томск", "Хабаровск", "Чита",
]

_DEFAULT_ROLE_IDS = [21, 31, 97, 102, 109]  # водитель, грузчик, продавец, разнорабочий, сварщик

_DEFAULT_SEARCH_TEXTS = ["Тракторист", "Монтер пути"]


def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    print("=" * 60)
    print("HH-Parser: Вакансии по гео и профессиям")
    print(f"Дата: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"Городов: {len(_DEFAULT_CITIES)}")
    print(f"Ролей: {len(_DEFAULT_ROLE_IDS)}, Текстовых запросов: {len(_DEFAULT_SEARCH_TEXTS)}")
    print("=" * 60)

    # 1. Резолвим города
    print("\n[1/3] Определяю area_id городов...")
    city_ids = resolve_city_ids(_DEFAULT_CITIES)
    print(f"  Найдено: {len(city_ids)}/{len(_DEFAULT_CITIES)} городов")

    if not city_ids:
        print("Нет городов для парсинга. Завершение.")
        return

    # 2. Парсим вакансии
    total_queries = len(city_ids) * (len(_DEFAULT_ROLE_IDS) + len(_DEFAULT_SEARCH_TEXTS))
    print(f"\n[2/3] Парсинг вакансий ({total_queries} запросов)...")

    def on_progress(p: dict) -> None:
        print(f"  {p['status_text']}", flush=True)

    result = scrape_vacancies_geo(
        city_ids=city_ids,
        role_ids=_DEFAULT_ROLE_IDS,
        search_texts=_DEFAULT_SEARCH_TEXTS,
        progress_cb=on_progress,
    )

    # 3. Сохраняем
    output_dir = Path(__file__).parent
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    output_path = str(output_dir / f"hh_vacancies_{timestamp}.xlsx")

    print(f"\n[3/3] Сохраняю результаты в {output_path}...")
    if HAS_OPENPYXL:
        export_to_xlsx(result, output_path)
        print(f"  Готово: {output_path}")
    else:
        print("  openpyxl не установлен, пропускаю экспорт в Excel")
        print("  Установите: pip install openpyxl")

    # 4. Итоговая статистика
    print("\n" + "=" * 60)
    print("ИТОГ — по текстовым запросам:")
    for text in _DEFAULT_SEARCH_TEXTS:
        rows = [r for r in result["texts"] if r["text"] == text and r["total"] >= 0]
        t = sum(r["total"] for r in rows)
        u = sum(r["unique"] for r in rows)
        print(f'  «{text}»: {t} всего, {u} уникальных')

    print("\nИТОГ — по ролям (сортировка по уникальным, все):")
    role_totals: dict[str, int] = {}
    role_unique: dict[str, int] = {}
    for r in result["roles"]:
        if r["total"] >= 0:
            role_totals[r["role_name"]] = role_totals.get(r["role_name"], 0) + r["total"]
            role_unique[r["role_name"]] = role_unique.get(r["role_name"], 0) + r["unique"]
    for role, u in sorted(role_unique.items(), key=lambda x: -x[1]):
        t = role_totals[role]
        print(f"  {role}: {t} всего, {u} уникальных")

    print("=" * 60)


if __name__ == "__main__":
    main()
