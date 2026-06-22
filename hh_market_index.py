"""
Блок 1 (ITD-1245): hh.индекс — рынок кандидатов.

hh.индекс = показатель напряжённости рынка (соотношение резюме/вакансий) по
профобласти в регионе. Источник — открытая статистика stats.hh.ru (см.
hh_market_stats.py). Берём последнее доступное значение (данные с лагом ~1 месяц).

Результат: сводная Excel «Город × Профобласть» со значением индекса.

Запуск из командной строки:
    python hh_market_index.py
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

import hh_market_stats as stats
from hh_vacancies_by_geo import load_area_tree, resolve_city_ids

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)

# Крупные города России с населением от ~500 тыс. (дефолтный набор для Блока 1).
# Можно переопределить списком при вызове build_index_table().
DEFAULT_CITIES_500K = [
    "Москва", "Санкт-Петербург", "Новосибирск", "Екатеринбург", "Казань",
    "Нижний Новгород", "Челябинск", "Красноярск", "Самара", "Уфа",
    "Ростов-на-Дону", "Краснодар", "Омск", "Воронеж", "Пермь",
    "Волгоград", "Саратов", "Тюмень", "Тольятти", "Ижевск",
    "Барнаул", "Ульяновск", "Иркутск", "Хабаровск", "Махачкала",
    "Владивосток", "Ярославль", "Оренбург", "Томск", "Кемерово",
    "Новокузнецк", "Рязань", "Набережные Челны", "Астрахань", "Пенза",
    "Липецк", "Киров", "Чебоксары", "Балашиха", "Калининград",
    "Тула", "Курск", "Севастополь",
]


def build_index_table(
    cities: Optional[list[str]] = None,
    prof_area_ids: Optional[list] = None,
    *,
    cache_path: Optional[str] = None,
    progress_cb: Optional[Callable[[dict], None]] = None,
) -> dict:
    """
    Собирает таблицу hh.индекса «город × профобласть» (последнее значение).

    Аргументы:
        cities        — список названий городов (по умолчанию DEFAULT_CITIES_500K).
        prof_area_ids — список id профобластей + опц. "all" (по умолчанию все +
                        агрегат "all" первым).
        cache_path    — путь к дисковому кэшу датасета stats.hh.ru.
        progress_cb   — коллбэк прогресса {done, total, status_text}.

    Возвращает dict:
        {
          "metric": "hhindex",
          "scraped_at": iso,
          "as_of": "май 2026",            # преобладающий период данных
          "cities": [город, ...],         # только найденные в датасете
          "prof_areas": [{"id","name"}],  # порядок столбцов
          "rows": [{city, area_id, prof_area_id, prof_area_name,
                    value, year, month, month_name}],
          "missing_cities": [...],        # не нашлись в дереве HH
          "missing_in_stats": [...],      # нашлись в дереве, но нет в датасете stats
        }
    """
    cities = cities or DEFAULT_CITIES_500K

    data = stats.fetch_stats_data(cache_path=cache_path)
    region_ids = stats.available_region_ids(data)
    name_map = stats.prof_area_name_map(include_all=True)

    # Профобласти-столбцы
    if prof_area_ids is None:
        prof_cols = stats.load_prof_areas(include_all=True)  # [{"id","name"}], "all" первым
    else:
        prof_cols = [{"id": pid, "name": name_map.get(str(pid), f"Профобласть {pid}")}
                     for pid in prof_area_ids]

    # Резолв городов -> area_id с fallback на родительский регион
    resolved = resolve_city_ids(cities)             # {city: area_id} (найденные в дереве)
    missing_cities = [c for c in cities if c not in resolved]
    areas_full = stats.load_areas_full()

    present: dict[str, dict] = {}     # city -> {"area_id","level","stats_name"}
    region_fallback: list[str] = []   # города, по которым взяли уровень региона
    missing_in_stats: list[str] = []
    for city, area_id in resolved.items():
        res = stats.resolve_to_stats_region(area_id, region_ids, areas_full=areas_full)
        if res is None:
            missing_in_stats.append(city)
            continue
        present[city] = {"area_id": res["id"], "level": res["level"], "stats_name": res["name"]}
        if res["level"] == "region":
            region_fallback.append(f'{city} → {res["name"]}')

    rows: list[dict] = []
    period_counter: dict[str, int] = {}
    total = len(present) * len(prof_cols)
    done = 0

    for city, meta in present.items():
        area_id = meta["area_id"]
        for pa in prof_cols:
            lv = stats.get_last(data, area_id, stats.METRIC_HH_INDEX, pa["id"])
            row = {
                "city": city,
                "area_id": area_id,
                "level": meta["level"],
                "stats_name": meta["stats_name"],
                "prof_area_id": pa["id"],
                "prof_area_name": pa["name"],
                "value": lv["value"] if lv else None,
                "year": lv["year"] if lv else None,
                "month": lv["month"] if lv else None,
                "month_name": lv["month_name"] if lv else None,
            }
            rows.append(row)
            if lv:
                key = f'{lv["month_name"]} {lv["year"]}'
                period_counter[key] = period_counter.get(key, 0) + 1
            done += 1
            if progress_cb:
                progress_cb({"done": done, "total": total,
                             "status_text": f'{city} — {pa["name"]}'})

    as_of = max(period_counter, key=period_counter.get) if period_counter else "—"

    return {
        "metric": stats.METRIC_HH_INDEX,
        "scraped_at": datetime.now().isoformat(timespec="seconds"),
        "as_of": as_of,
        "cities": list(present.keys()),
        "prof_areas": prof_cols,
        "rows": rows,
        "region_fallback": region_fallback,
        "missing_cities": missing_cities,
        "missing_in_stats": missing_in_stats,
    }


def export_index_to_xlsx(result: dict, output_path: str) -> None:
    """
    Экспорт результата build_index_table() в Excel: сводная «Город × Профобласть».

    Значение в ячейке — последнее доступное hh.индекс. Над таблицей — пометка о
    периоде данных и источнике. Заголовки выделены, числа выровнены по центру.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("Нужен openpyxl: pip install openpyxl")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    note_font = Font(italic=True, color="808080")
    city_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "hh.индекс"

    ws.append(["hh.индекс (резюме/вакансий) — рынок кандидатов"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=13)
    ws.append([f"Источник: stats.hh.ru · данные на: {result.get('as_of', '—')} · "
               f"выгружено: {result.get('scraped_at', '')}"])
    ws.cell(ws.max_row, 1).font = note_font
    ws.append([])

    prof_areas = result["prof_areas"]
    cities = result["cities"]

    header = ["Город / Регион"] + [pa["name"] for pa in prof_areas]
    ws.append(header)
    hr = ws.max_row
    for cell in ws[hr]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Индекс по (city, prof_area_id) -> value; уровень данных по городу
    idx = {(r["city"], str(r["prof_area_id"])): r["value"] for r in result["rows"]}
    level_by_city = {r["city"]: r.get("level", "city") for r in result["rows"]}

    for city in cities:
        # Города с данными уровня региона помечаем «*»
        label = f"{city} *" if level_by_city.get(city) == "region" else city
        line = [label]
        for pa in prof_areas:
            v = idx.get((city, str(pa["id"])))
            line.append(v if v is not None else "—")
        ws.append(line)
        ws.cell(ws.max_row, 1).font = city_font
        for c in range(2, len(header) + 1):
            ws.cell(ws.max_row, c).alignment = center

    # Пометки о пропусках и fallback
    if result.get("region_fallback") or result.get("missing_cities") or result.get("missing_in_stats"):
        ws.append([])
        if result.get("region_fallback"):
            ws.append(["* данные уровня региона (города нет в статистике отдельно): "
                       + "; ".join(result["region_fallback"])])
            ws.cell(ws.max_row, 1).font = note_font
        if result.get("missing_cities"):
            ws.append([f"Не найдены в справочнике HH: {', '.join(result['missing_cities'])}"])
            ws.cell(ws.max_row, 1).font = note_font
        if result.get("missing_in_stats"):
            ws.append([f"Нет в статистике stats.hh.ru: {', '.join(result['missing_in_stats'])}"])
            ws.cell(ws.max_row, 1).font = note_font

    # Ширина колонок
    ws.column_dimensions["A"].width = 22
    for i in range(2, len(header) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 16

    ws.freeze_panes = ws.cell(row=hr + 1, column=2)
    wb.save(output_path)
    logger.info("Сохранено: %s", output_path)


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s",
                        datefmt="%H:%M:%S")
    cache = str(Path(__file__).parent / "data" / "stats_hh_cache.json")

    print("=" * 60)
    print("Блок 1: hh.индекс (рынок кандидатов)")
    print(f"Городов: {len(DEFAULT_CITIES_500K)}, профобластей: все + агрегат")
    print("=" * 60)

    result = build_index_table(
        cache_path=cache,
        progress_cb=lambda p: print(f"  [{p['done']}/{p['total']}] {p['status_text']}", flush=True),
    )

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out = str(Path(__file__).parent / f"hh_index_{ts}.xlsx")
    export_index_to_xlsx(result, out)
    print(f"\nГотово: {out}")
    print(f"Города: {len(result['cities'])}, профобласти: {len(result['prof_areas'])}, "
          f"данные на: {result['as_of']}")
    if result["missing_in_stats"]:
        print(f"Нет в статистике: {', '.join(result['missing_in_stats'])}")
    if result["missing_cities"]:
        print(f"Не найдены в HH: {', '.join(result['missing_cities'])}")


if __name__ == "__main__":
    main()
